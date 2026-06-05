import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import MergedCell
from copy import copy
from datetime import datetime, timedelta
from io import BytesIO
import os
import re
import gc
import shutil
import tempfile
import traceback
import warnings
import unicodedata

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

st.set_page_config(page_title="Consolidador Gerencial CN", page_icon="📊", layout="wide")

EXTENSOES_EXCEL_VALIDAS = (".xlsx", ".xlsm")
NOME_ABA_GERENCIAL = "Gerencial"
COLUNA_INICIO_LIMPEZA = 13  # Coluna M


def remover_acentos(texto):
    texto = str(texto)
    texto = unicodedata.normalize("NFKD", texto)
    texto = texto.encode("ASCII", "ignore").decode("ASCII")
    return texto


def limpar_nome_aba(nome):
    nome = os.path.splitext(str(nome))[0]
    nome = re.sub(r'[:\\/?*\[\]]', '', nome).strip()
    return (nome if nome else "Aba")[:31]


def gerar_nome_aba_unico(nome_base, nomes_existentes):
    nome_base = limpar_nome_aba(nome_base)
    if nome_base not in nomes_existentes:
        return nome_base
    contador = 1
    while True:
        sufixo = f" ({contador})"
        nome_tentativa = nome_base[:31 - len(sufixo)] + sufixo
        if nome_tentativa not in nomes_existentes:
            return nome_tentativa
        contador += 1


def limpar_nome_arquivo(nome):
    nome = re.sub(r'[\\/:*?"<>|]', '', str(nome)).strip()
    return nome if nome else "arquivo.xlsx"


def normalizar_texto(texto):
    return str(texto).strip().lower()


def normalizar_nome_para_mapeamento(nome):
    nome = os.path.splitext(str(nome))[0]
    nome = remover_acentos(nome).lower()
    nome = re.sub(r"[_\-\.]+", " ", nome)
    nome = re.sub(r"\s+", " ", nome)
    return nome.strip()


def obter_nome_aba_final_personalizado(nome_arquivo_origem):
    """
    Regras de nome final:
    - Gerencial_XX, onde XX é dia variável -> XAM
    - Gerencial DD-MM-AAAA -> COB
    - Gerencial XX, com espaço e dia variável -> PVE
    - Resumo Gerencial Diário - EDE - -> EDE
    - RG Sobradinho -> SOB
    - Resumo Gerencial Cuiabá -> CUI
    - Resumo Gerencial genérico -> NOB
    """
    nome_original_sem_ext = os.path.splitext(str(nome_arquivo_origem))[0]
    nome_sem_acento = remover_acentos(nome_original_sem_ext).lower().strip()
    nome_sem_acento = re.sub(r"\s+", " ", nome_sem_acento)
    nome_normalizado = normalizar_nome_para_mapeamento(nome_arquivo_origem)

    if "resumo gerencial cuiaba" in nome_normalizado:
        return "CUI"
    if "resumo gerencial diario ede" in nome_normalizado:
        return "EDE"
    if "rg sobradinho" in nome_normalizado:
        return "SOB"

    # XAM: Gerencial_XX, com underline e dia variável. Ex.: Gerencial_04, Gerencial_15.
    if re.search(r"(^|[^a-z0-9])gerencial_([0-2]?\d|3[01])([^0-9]|$)", nome_sem_acento):
        return "XAM"

    # COB: Gerencial DD-MM-AAAA, com espaço e data completa com hífen.
    if re.search(r"\bgerencial\s+([0-2]\d|3[01])-[01]\d-\d{4}\b", nome_sem_acento):
        return "COB"

    # PVE: Gerencial XX, com espaço e dia variável, sem hífen logo após o dia.
    if re.search(r"\bgerencial\s+([0-2]?\d|3[01])([^0-9-]|$)", nome_sem_acento):
        return "PVE"

    if "resumo gerencial" in nome_normalizado:
        return "NOB"

    return limpar_nome_aba(nome_arquivo_origem)


def interpretar_aba_como_data(nome_aba, ano_referencia, mes_referencia):
    nome = str(nome_aba).strip()
    if not nome.isdigit():
        return None
    try:
        if len(nome) in [1, 2]:
            return datetime(ano_referencia, mes_referencia, int(nome))
        if len(nome) == 4:
            return datetime(ano_referencia, int(nome[2:]), int(nome[:2]))
        return None
    except ValueError:
        return None


def listar_datas_disponiveis_arquivos(arquivos_salvos, ano_referencia, mes_referencia, quantidade_datas_filtro=5):
    datas_encontradas = {}
    for item in arquivos_salvos:
        wb = None
        try:
            caminho = item["caminho"]
            if not os.path.exists(caminho):
                continue
            wb = load_workbook(caminho, read_only=True, data_only=True, keep_links=False)
            for nome_aba in wb.sheetnames:
                data_interpretada = interpretar_aba_como_data(nome_aba, ano_referencia, mes_referencia)
                if data_interpretada is not None:
                    chave = data_interpretada.date()
                    datas_encontradas.setdefault(chave, {"data": data_interpretada, "abas": [], "arquivos": []})
                    datas_encontradas[chave]["abas"].append(nome_aba)
                    datas_encontradas[chave]["arquivos"].append(item["nome"])
        except Exception:
            pass
        finally:
            try:
                if wb:
                    wb.close()
            except Exception:
                pass
            gc.collect()

    if not datas_encontradas:
        return []

    lista_datas = [
        item for item in datas_encontradas.values()
        if item["data"].year == ano_referencia and item["data"].month == mes_referencia
    ]
    if not lista_datas:
        return []

    lista_datas = sorted(lista_datas, key=lambda item: item["data"], reverse=True)
    sequencia = []
    data_anterior = None
    for item in lista_datas:
        data_atual = item["data"].date()
        if data_anterior is None:
            sequencia.append(item)
            data_anterior = data_atual
        elif (data_anterior - data_atual).days == 1:
            sequencia.append(item)
            data_anterior = data_atual
        else:
            break
        if len(sequencia) >= quantidade_datas_filtro:
            break
    return sequencia


def encontrar_aba_gerencial(wb_origem):
    for nome_aba in wb_origem.sheetnames:
        if normalizar_texto(nome_aba) == normalizar_texto(NOME_ABA_GERENCIAL):
            return nome_aba
    return None


def encontrar_aba_por_data_ou_gerencial(wb_origem, data_selecionada, ano_referencia, mes_referencia):
    abas_compativeis = []
    for nome_aba in wb_origem.sheetnames:
        data_interpretada = interpretar_aba_como_data(nome_aba, ano_referencia, mes_referencia)
        if data_interpretada is not None and data_interpretada.date() == data_selecionada:
            nome = str(nome_aba).strip()
            if len(nome) == 4:
                prioridade = 1
            elif len(nome) == 2:
                prioridade = 2
            elif len(nome) == 1:
                prioridade = 3
            else:
                prioridade = 9
            abas_compativeis.append({"nome_aba": nome_aba, "prioridade": prioridade})

    if abas_compativeis:
        abas_compativeis = sorted(abas_compativeis, key=lambda item: item["prioridade"])
        return abas_compativeis[0]["nome_aba"], "Aba da data selecionada"

    aba_gerencial = encontrar_aba_gerencial(wb_origem)
    if aba_gerencial:
        return aba_gerencial, "Aba Gerencial usada porque a aba da data selecionada não foi encontrada"

    return None, "Nenhuma aba da data selecionada ou aba Gerencial encontrada"


def criar_pasta_sessao():
    if "pasta_sessao" not in st.session_state:
        st.session_state.pasta_sessao = tempfile.mkdtemp()


def salvar_uploads_em_pasta_sessao(arquivos_upload):
    criar_pasta_sessao()
    arquivos_salvos = []
    for uploaded_file in arquivos_upload:
        nome_limpo = limpar_nome_arquivo(uploaded_file.name)
        caminho_base = os.path.join(st.session_state.pasta_sessao, nome_limpo)
        caminho_final = caminho_base
        contador = 1
        while os.path.exists(caminho_final):
            nome_sem_extensao, extensao = os.path.splitext(nome_limpo)
            caminho_final = os.path.join(st.session_state.pasta_sessao, f"{nome_sem_extensao}_{contador}{extensao}")
            contador += 1
        with open(caminho_final, "wb") as f:
            f.write(uploaded_file.getvalue())
        arquivos_salvos.append({"nome": uploaded_file.name, "caminho": caminho_final, "tamanho": uploaded_file.size})
    return arquivos_salvos


def limpar_arquivos_da_sessao():
    if "pasta_sessao" in st.session_state:
        try:
            if os.path.exists(st.session_state.pasta_sessao):
                shutil.rmtree(st.session_state.pasta_sessao)
        except Exception:
            pass
    st.session_state.arquivos_salvos = []
    st.session_state.uploader_key += 1
    if "pasta_sessao" in st.session_state:
        del st.session_state.pasta_sessao


def aplicar_configuracoes_finais_aba(aba_destino):
    aba_destino.sheet_view.zoomScale = 80
    aba_destino.sheet_view.zoomScaleNormal = 80
    aba_destino.freeze_panes = "A5"
    aba_destino.sheet_view.showGridLines = False


def limpar_conteudos_a_partir_da_coluna_m(aba_destino):
    """
    Remove valores/conteúdos da coluna M em diante, preservando estrutura e formatação.
    Coluna M = índice 13.
    Também ignora células mescladas não editáveis para evitar erro em abas como PVE.
    """
    max_linha = aba_destino.max_row
    max_coluna = aba_destino.max_column

    if max_coluna < COLUNA_INICIO_LIMPEZA:
        return

    for linha in aba_destino.iter_rows(
        min_row=1,
        max_row=max_linha,
        min_col=COLUNA_INICIO_LIMPEZA,
        max_col=max_coluna
    ):
        for celula in linha:
            if isinstance(celula, MergedCell):
                continue
            celula.value = None
            celula.comment = None
            celula._hyperlink = None


def padronizar_larguras_colunas_pela_aba_ede(wb_final, nome_aba_referencia="EDE"):
    if nome_aba_referencia not in wb_final.sheetnames:
        return False, f"A aba de referência '{nome_aba_referencia}' não foi encontrada. As larguras das colunas não foram padronizadas."
    aba_referencia = wb_final[nome_aba_referencia]
    max_colunas = max(ws.max_column for ws in wb_final.worksheets)
    for indice_coluna in range(1, max_colunas + 1):
        letra_coluna = aba_referencia.cell(row=1, column=indice_coluna).column_letter
        largura_referencia = aba_referencia.column_dimensions[letra_coluna].width
        if largura_referencia is None:
            largura_referencia = aba_referencia.sheet_format.defaultColWidth or 8.43
        for aba in wb_final.worksheets:
            aba.column_dimensions[letra_coluna].width = largura_referencia
    return True, f"Larguras das colunas padronizadas com base na aba '{nome_aba_referencia}'."


def copiar_aba_como_valores_mais_leve(aba_origem, aba_destino, copiar_estilos=True):
    for linha in aba_origem.iter_rows(min_row=1, max_row=aba_origem.max_row, min_col=1, max_col=aba_origem.max_column):
        for celula_origem in linha:
            celula_destino = aba_destino[celula_origem.coordinate]
            celula_destino.value = celula_origem.value
            if copiar_estilos and celula_origem.has_style:
                celula_destino.font = copy(celula_origem.font)
                celula_destino.fill = copy(celula_origem.fill)
                celula_destino.border = copy(celula_origem.border)
                celula_destino.alignment = copy(celula_origem.alignment)
                celula_destino.number_format = celula_origem.number_format
                celula_destino.protection = copy(celula_origem.protection)
            if celula_origem.comment:
                celula_destino.comment = copy(celula_origem.comment)
            if celula_origem.hyperlink:
                celula_destino._hyperlink = copy(celula_origem.hyperlink)

    for letra_coluna, dimensao_coluna in aba_origem.column_dimensions.items():
        aba_destino.column_dimensions[letra_coluna].width = dimensao_coluna.width
        aba_destino.column_dimensions[letra_coluna].hidden = dimensao_coluna.hidden
        aba_destino.column_dimensions[letra_coluna].outlineLevel = dimensao_coluna.outlineLevel
        aba_destino.column_dimensions[letra_coluna].collapsed = dimensao_coluna.collapsed

    for numero_linha, dimensao_linha in aba_origem.row_dimensions.items():
        aba_destino.row_dimensions[numero_linha].height = dimensao_linha.height
        aba_destino.row_dimensions[numero_linha].hidden = dimensao_linha.hidden
        aba_destino.row_dimensions[numero_linha].outlineLevel = dimensao_linha.outlineLevel
        aba_destino.row_dimensions[numero_linha].collapsed = dimensao_linha.collapsed

    for intervalo_mesclado in aba_origem.merged_cells.ranges:
        aba_destino.merge_cells(str(intervalo_mesclado))

    if aba_origem.auto_filter and aba_origem.auto_filter.ref:
        aba_destino.auto_filter.ref = aba_origem.auto_filter.ref

    aba_destino.freeze_panes = aba_origem.freeze_panes

    try:
        aba_destino.sheet_view.showGridLines = aba_origem.sheet_view.showGridLines
    except Exception:
        pass

    try:
        aba_destino.page_setup.orientation = aba_origem.page_setup.orientation
        aba_destino.page_setup.paperSize = aba_origem.page_setup.paperSize
        aba_destino.page_setup.fitToWidth = aba_origem.page_setup.fitToWidth
        aba_destino.page_setup.fitToHeight = aba_origem.page_setup.fitToHeight
        aba_destino.page_margins.left = aba_origem.page_margins.left
        aba_destino.page_margins.right = aba_origem.page_margins.right
        aba_destino.page_margins.top = aba_origem.page_margins.top
        aba_destino.page_margins.bottom = aba_origem.page_margins.bottom
        aba_destino.page_margins.header = aba_origem.page_margins.header
        aba_destino.page_margins.footer = aba_origem.page_margins.footer
    except Exception:
        pass

    try:
        if aba_origem.print_area:
            aba_destino.print_area = aba_origem.print_area
    except Exception:
        pass

    try:
        aba_destino.print_title_rows = aba_origem.print_title_rows
        aba_destino.print_title_cols = aba_origem.print_title_cols
    except Exception:
        pass


def processar_arquivos_salvos(arquivos_salvos, data_referencia, data_selecionada, copiar_estilos=True):
    logs = []
    arquivos_processados = []
    arquivos_com_erro = []
    abas_criadas = []
    ano_referencia = data_referencia.year
    mes_referencia = data_referencia.month
    nome_arquivo_final = f"Resumo Gerencial CN - {data_selecionada.strftime('%d.%m.%Y')}.xlsx"
    wb_final = Workbook()
    wb_final.remove(wb_final.active)

    logs.append("Iniciando processamento.")
    logs.append(f"Data selecionada no filtro: {data_selecionada.strftime('%d/%m/%Y')}")
    logs.append(f"Arquivos carregados: {len(arquivos_salvos)}")
    logs.append("Regra: copia a aba da data selecionada; se não existir, copia a aba Gerencial.")
    logs.append("Configuração final: zoom 80%, congelamento em A5, linhas de grade ocultas, larguras padronizadas pela aba EDE e conteúdos removidos da coluna M em diante.")

    barra_progresso = st.progress(0)
    texto_status = st.empty()

    for indice, item in enumerate(arquivos_salvos, start=1):
        nome_original = item["nome"]
        caminho_arquivo = item["caminho"]
        wb_origem = None
        texto_status.info(f"Processando {indice}/{len(arquivos_salvos)}: {nome_original}")
        logs.append("")
        logs.append(f"[{indice}/{len(arquivos_salvos)}] Processando: {nome_original}")
        try:
            if not nome_original.lower().endswith(EXTENSOES_EXCEL_VALIDAS):
                raise ValueError("Extensão inválida. Use .xlsx ou .xlsm.")
            if not os.path.exists(caminho_arquivo):
                raise FileNotFoundError("Arquivo temporário não encontrado. Carregue o arquivo novamente.")
            wb_origem = load_workbook(caminho_arquivo, data_only=True, keep_links=False)
            nome_aba_origem, criterio_aba = encontrar_aba_por_data_ou_gerencial(wb_origem, data_selecionada, ano_referencia, mes_referencia)
            if not nome_aba_origem:
                abas_disponiveis = ", ".join(wb_origem.sheetnames)
                raise ValueError(f"Nenhuma aba correspondente à data {data_selecionada.strftime('%d/%m/%Y')} ou aba Gerencial foi encontrada. Abas disponíveis: {abas_disponiveis}")

            aba_origem = wb_origem[nome_aba_origem]
            nome_base_personalizado = obter_nome_aba_final_personalizado(nome_original)
            nome_aba_final = gerar_nome_aba_unico(nome_base_personalizado, wb_final.sheetnames)
            aba_destino = wb_final.create_sheet(title=nome_aba_final)

            copiar_aba_como_valores_mais_leve(aba_origem, aba_destino, copiar_estilos=copiar_estilos)
            limpar_conteudos_a_partir_da_coluna_m(aba_destino)
            aplicar_configuracoes_finais_aba(aba_destino)

            arquivos_processados.append({
                "arquivo": nome_original,
                "aba_origem": nome_aba_origem,
                "aba_final": nome_aba_final,
                "criterio": f"{criterio_aba} | Data selecionada no filtro: {data_selecionada.strftime('%d/%m/%Y')}"
            })
            abas_criadas.append(nome_aba_final)
            logs.append(f"SUCESSO: aba '{nome_aba_origem}' copiada para '{nome_aba_final}'.")
            logs.append(f"Critério usado: {criterio_aba}")
            logs.append("Configurações aplicadas: zoom 80%, freeze panes em A5, sem linhas de grade e conteúdos limpos da coluna M em diante.")
        except Exception as erro:
            arquivos_com_erro.append({"arquivo": nome_original, "erro": str(erro)})
            logs.append(f"ERRO: {nome_original} - {erro}")
        finally:
            try:
                if wb_origem:
                    wb_origem.close()
            except Exception:
                pass
            try:
                del wb_origem
            except Exception:
                pass
            gc.collect()
            barra_progresso.progress(indice / len(arquivos_salvos))

    texto_status.success("Processamento concluído.")

    if len(wb_final.sheetnames) == 0:
        raise Exception("Nenhuma aba foi criada no arquivo final.")

    _, mensagem_larguras = padronizar_larguras_colunas_pela_aba_ede(wb_final, "EDE")
    logs.append(mensagem_larguras)

    caminho_saida_temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
    wb_final.save(caminho_saida_temp)
    wb_final.close()

    with open(caminho_saida_temp, "rb") as f:
        conteudo_final = f.read()

    try:
        os.remove(caminho_saida_temp)
    except Exception:
        pass

    output = BytesIO(conteudo_final)
    output.seek(0)
    logs.append("")
    logs.append("Arquivo final gerado com sucesso.")
    logs.append(f"Nome do arquivo: {nome_arquivo_final}")
    gc.collect()

    return {
        "arquivo_excel": output,
        "nome_arquivo_final": nome_arquivo_final,
        "logs": logs,
        "arquivos_processados": arquivos_processados,
        "arquivos_com_erro": arquivos_com_erro,
        "abas_criadas": abas_criadas
    }


if "arquivos_salvos" not in st.session_state:
    st.session_state.arquivos_salvos = []
if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = 0

st.title("📊 Sistema de Consolidação Gerencial CN")
st.markdown(
    """
    Este sistema consolida abas de arquivos Excel em um único arquivo final.

    **Como funciona:**

    - Você envia todos os arquivos Excel de uma vez;
    - O sistema lê as abas numéricas dos arquivos;
    - O filtro mostra as **últimas datas em sequência contínua** dentro do mesmo mês;
    - Você escolhe **uma data**;
    - O sistema tenta copiar a aba correspondente à data escolhida;
    - Se um arquivo não tiver a aba da data, o sistema copia a aba **Gerencial**;
    - O arquivo final é gerado com nomes padronizados: **PVE, COB, NOB, EDE, SOB, XAM, CUI**;
    - Todas as abas finais ficam com **zoom 80%**, **painéis congelados em A5**, **sem linhas de grade**, **larguras de colunas iguais às da aba EDE** e **sem conteúdos da coluna M em diante**.
    """
)

st.divider()

st.sidebar.header("⚙️ Configurações")
usar_data_ontem = st.sidebar.checkbox("Usar mês/ano de ontem como referência", value=True)
if usar_data_ontem:
    data_referencia = datetime.today() - timedelta(days=1)
    st.sidebar.info(f"Mês/ano de referência: {data_referencia.strftime('%m/%Y')}")
else:
    data_escolhida = st.sidebar.date_input("Escolha uma data de referência para mês/ano", value=(datetime.today() - timedelta(days=1)).date())
    data_referencia = datetime.combine(data_escolhida, datetime.min.time())

qtde_esperada = st.sidebar.number_input("Quantidade esperada de arquivos", min_value=1, max_value=100, value=7, step=1)
copiar_estilos = st.sidebar.checkbox("Copiar formatação das células", value=True)
quantidade_datas_filtro = st.sidebar.number_input("Quantidade máxima de datas no filtro", min_value=1, max_value=31, value=5, step=1)
st.sidebar.caption("O filtro mostra somente datas consecutivas dentro do mês/ano de referência. Exemplo: 06, 05, 04, 03, 02. Não mistura 29, 30, 31 com 01, 02 de outro mês.")

st.subheader("1. Envie os arquivos Excel")
arquivos_upload = st.file_uploader("Selecione todos os arquivos Excel de uma vez", type=["xlsx", "xlsm"], accept_multiple_files=True, key=f"uploader_{st.session_state.uploader_key}")

col_carregar, col_limpar = st.columns(2)
with col_carregar:
    if st.button("📥 Carregar arquivos enviados", disabled=not arquivos_upload):
        try:
            limpar_arquivos_da_sessao()
            st.session_state.arquivos_salvos = salvar_uploads_em_pasta_sessao(arquivos_upload)
            st.success(f"{len(st.session_state.arquivos_salvos)} arquivo(s) carregado(s) com sucesso.")
            st.rerun()
        except Exception as erro:
            st.error(f"Erro ao carregar arquivos: {erro}")
with col_limpar:
    if st.button("🧹 Limpar arquivos carregados"):
        limpar_arquivos_da_sessao()
        st.success("Arquivos carregados foram limpos.")
        st.rerun()

arquivos_salvos = st.session_state.arquivos_salvos
if arquivos_salvos:
    st.success(f"{len(arquivos_salvos)} arquivo(s) carregado(s).")
    if len(arquivos_salvos) != qtde_esperada:
        st.warning(f"Foram carregados {len(arquivos_salvos)} arquivo(s), mas a quantidade esperada é {qtde_esperada}.")
    tamanho_total_mb = sum(item["tamanho"] for item in arquivos_salvos) / (1024 * 1024)
    st.info(f"Tamanho total carregado: {tamanho_total_mb:.2f} MB")
    if tamanho_total_mb > 30:
        st.warning("Os arquivos carregados somam mais de 30 MB. Se o sistema cair, tente desmarcar a opção de copiar formatação.")
    with st.expander("Ver arquivos carregados"):
        for i, item in enumerate(arquivos_salvos, start=1):
            tamanho_mb = item["tamanho"] / (1024 * 1024)
            nome_padronizado = obter_nome_aba_final_personalizado(item["nome"])
            st.write(f"{i}. {item['nome']} — {tamanho_mb:.2f} MB — aba final: {nome_padronizado}")
else:
    st.info("Envie os arquivos Excel e clique em 'Carregar arquivos enviados'.")

st.divider()
st.subheader("2. Escolha a data da aba que será copiada")
data_selecionada = None
if arquivos_salvos:
    datas_para_filtro = listar_datas_disponiveis_arquivos(arquivos_salvos, data_referencia.year, data_referencia.month, quantidade_datas_filtro)
    if datas_para_filtro:
        opcoes_datas = []
        for item in datas_para_filtro:
            data_item = item["data"].date()
            qtd_arquivos = len(set(item["arquivos"]))
            label = f"{data_item.strftime('%d/%m/%Y')} — encontrada em {qtd_arquivos} arquivo(s)"
            opcoes_datas.append({"label": label, "data": data_item})
        opcao_escolhida = st.selectbox("Selecione uma data sequencial encontrada nas abas", options=opcoes_datas, format_func=lambda item: item["label"])
        data_selecionada = opcao_escolhida["data"]
        st.success(f"Data selecionada: {data_selecionada.strftime('%d/%m/%Y')}")
        st.info("Regra de cópia: o sistema tenta copiar a aba da data selecionada. Se não encontrar essa aba em algum arquivo, copia a aba Gerencial. Depois aplica zoom 80%, congela em A5, oculta linhas de grade, padroniza larguras pela aba EDE e limpa conteúdos da coluna M em diante.")
        with st.expander("Datas exibidas no filtro"):
            for opcao in opcoes_datas:
                st.write(f"- {opcao['label']}")
    else:
        st.warning("Nenhuma sequência de datas foi encontrada no mês/ano de referência. Verifique se as abas estão como 1, 01 ou 0106 e se o mês/ano de referência está correto. Se os arquivos tiverem apenas aba Gerencial, adicione pelo menos um arquivo com aba de data para aparecer no filtro.")
else:
    st.info("Carregue os arquivos para o sistema listar as datas disponíveis.")

st.divider()
st.subheader("3. Processar arquivos")
botao_processar = st.button("🚀 Processar data selecionada", type="primary", disabled=len(st.session_state.arquivos_salvos) == 0 or data_selecionada is None)
if botao_processar:
    try:
        with st.spinner("Processando arquivos..."):
            resultado = processar_arquivos_salvos(st.session_state.arquivos_salvos, data_referencia, data_selecionada, copiar_estilos)
        st.success("Arquivo final gerado com sucesso!")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Arquivos carregados", len(st.session_state.arquivos_salvos))
        with col2:
            st.metric("Processados com sucesso", len(resultado["arquivos_processados"]))
        with col3:
            st.metric("Com erro", len(resultado["arquivos_com_erro"]))
        st.subheader("4. Baixar arquivo final")
        st.download_button(label="⬇️ Baixar Excel consolidado", data=resultado["arquivo_excel"], file_name=resultado["nome_arquivo_final"], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.subheader("5. Resumo")
        if resultado["arquivos_processados"]:
            st.markdown("### ✅ Arquivos processados")
            for item in resultado["arquivos_processados"]:
                st.write(f"- **Arquivo:** {item['arquivo']} | **Aba origem:** {item['aba_origem']} | **Aba final:** {item['aba_final']} | **Critério:** {item['criterio']}")
        if resultado["abas_criadas"]:
            with st.expander("Abas criadas no arquivo final"):
                for aba in resultado["abas_criadas"]:
                    st.write(f"- {aba}")
        if resultado["arquivos_com_erro"]:
            st.markdown("### ❌ Arquivos com erro")
            for item in resultado["arquivos_com_erro"]:
                st.error(f"{item['arquivo']}: {item['erro']}")
        with st.expander("Ver logs detalhados"):
            st.code("\n".join(resultado["logs"]), language="text")
    except Exception as erro:
        st.error(f"Erro geral ao processar os arquivos: {erro}")
        st.code(traceback.format_exc(), language="text")
        st.warning("Se o app cair ou travar com vários arquivos, tente desmarcar 'Copiar formatação das células' na barra lateral.")

st.divider()
st.caption("Observação: o sistema copia a aba da data selecionada. Se algum arquivo não tiver essa aba, o sistema usa a aba Gerencial como alternativa. Todas as abas finais ficam com zoom 80%, painéis congelados em A5, linhas de grade ocultas, larguras de colunas iguais às da aba EDE e sem conteúdos da coluna M em diante. O filtro considera somente datas consecutivas no mesmo mês/ano de referência.")
